'''
This program is meant to take help with the input of Products into the ACES file to uplaod to SEMAdata
Requires: User Input (Loading List) , Jobber Sheet , and ACES Sheet
Author = Edmond Wu 
''' 
import tkinter as tk
from tkinter import messagebox
import openpyxl 
from openpyxl import load_workbook


''' GUI For Tool
Take in Loading List through User Input GUI 
Creates a file called loading_list 
Press Enter to add to loading_list
Press Review to check loading_list
Press Done when input is done
'''

def save_to_file(event=None): # Save input to file 
    user_input = input_field.get()
    if user_input.strip():
        with open("loading_list.txt", "a") as file:
            file.write(user_input + "\n")
        input_field.delete(0, tk.END)

def close_gui(): #Close GUI 
    root.destroy()

def review_file(): #Display Contents of file in MessageBox
    with open("loading_list.txt", "r") as file:
        contents = file.read()
        messagebox.showinfo("File Contents", contents)

root = tk.Tk()
root.title("Loading List Input")

input_label = tk.Label(root, text="Enter Part Number:")
input_label.pack()

input_field = tk.Entry(root)
input_field.pack(padx=50 , pady=50)

input_field.bind("<Return>", save_to_file)

done_button = tk.Button(root, text="Done", command=close_gui)
done_button.pack(side=tk.LEFT , fill=tk.X , expand=True)

review_button = tk.Button(root, text="Review", command=review_file)
review_button.pack(side=tk.LEFT , fill=tk.X , expand=True)

#root.mainloop()

# Open excel files for reading 
# Maybe add a function to give the option to upload a file 
jobber_book = load_workbook('jobber.xlsx')
#aces_book = load_workbook('aces.xlsm') This shit takes forever to load 
jobber_sheet = jobber_book.active
all_part_num = [] #Initialize empty list for storing part num

#Reading the jobber file 3rd column
for row in jobber_sheet.iter_cols(min_col=3, max_col=3):
    for cell in row:
        all_part_num.append(cell.value)
del all_part_num [0:3]

#Reads the loading list into a list 
loading_list = open("loading_list.txt" , "r")
loading_list_parts = loading_list.readlines()
loading_list_parts = [element.strip() for element in loading_list_parts] # Removes \n


#Filters all_part_num using loading_list_parts and outputs filter parts into filter_part.txt
with open("filtered_parts.txt", "a") as output_file:
    for number in loading_list_parts:
        filtered_items = [item for item in all_part_num if number in item]
        if filtered_items:
            output_file.write(f"Part Number : {number}\n")
            for item in filtered_items:
                output_file.write(f"- {item}\n")
            output_file.write("\n")
        else:
            print(f"No items found for part {number}")

